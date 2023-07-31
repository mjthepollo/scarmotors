from datetime import date

from dateutil.relativedelta import relativedelta
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from core.utility import print_colored
from demand.excel_line_info import *
from demand.sales_models import Order, Register

# Create a logger with a custom log level


class NoDateException(Exception):
    pass


def get_incentive_commnet(row):
    return row[SUPPORTER].comment.text.replace("Windows 사용자:", "").replace("\n", "")


def get_merged_cell_value(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]

    cell_value = sheet.cell(rng[0].min_row, rng[0].min_col).value if len(
        rng) != 0 else cell.value

    if len(rng) != 0:
        index = list(rng[0].cols)[0].index((cell.row, cell.column))
        return cell_value, index
    else:
        return cell_value, 0


def get_incentive_date_from_row(row):
    comment = get_incentive_commnet(row)
    if "1월" in comment:
        raw_date = date(2023, 1, 1)
    elif "2월" in comment:
        raw_date = date(2023, 2, 1)
    elif "3월" in comment:
        raw_date = date(2023, 3, 1)
    elif "4월" in comment:
        raw_date = date(2023, 4, 1)
    elif "5월" in comment:
        raw_date = date(2023, 5, 1)
    elif "5얼" in comment:
        raw_date = date(2023, 5, 1)
    elif "6월" in comment:
        raw_date = date(2023, 6, 1)
    elif "6얼" in comment:
        raw_date = date(2023, 6, 1)
    elif "7월" in comment:
        raw_date = date(2023, 7, 1)
    elif "8월" in comment:
        raw_date = date(2022, 8, 1)
    elif "9월" in comment:
        raw_date = date(2022, 9, 1)
    elif "10월" in comment:
        raw_date = date(2022, 10, 1)
    elif "11월" in comment:
        raw_date = date(2022, 11, 1)
    elif "12월" in comment:
        raw_date = date(2022, 12, 1)
    else:
        print(f"WTF!! {comment}")
        raise NoDateException()
    return raw_date + relativedelta(months=1)


workbook = load_workbook('src/data_load.xlsx')
sheet_names = ["22년 12월 미청구", "23년 본사 상반기", "23년 본사 하반기"]
INCENTIVED_SET = set()
NOT_INCENTIVED_SET = set()


def get_incentive_data(sheet_index, row_index, row, date):
    return f"{row[RO_NUMBER].value}/{row[RECEIPT_NUMBER].value}/{row[CAR_NUMBER].value}/{row[SUPPORTER].value}:{date.strftime('%y/%m')}"


class Command(BaseCommand):
    help = 'clean models by 입출고대장'

    def handle(self, *args, **options):
        for sheet_index, worksheet in enumerate(sheet_names):
            worksheet = workbook.get_sheet_by_name(worksheet)
            for row_index, row in enumerate(worksheet.iter_rows()):
                if row[SUPPORTER].comment:
                    try:
                        date = get_incentive_date_from_row(row)
                    except NoDateException:
                        print_colored(
                            f"NO DATE: {row[RECEIPT_NUMBER].value}/{row[CAR_NUMBER].value}", "yellow")
                        continue
                    receipt_number = row[RECEIPT_NUMBER].value
                    RO_number, index = get_merged_cell_value(
                        worksheet, row[RO_NUMBER])
                    try:
                        register = Register.objects.get(RO_number=RO_number)
                    except Register.DoesNotExist as e:
                        print_colored(
                            f"WORK SHEET : {sheet_names[sheet_index]}, ROW_INDEX:{row_index}", "yellow")
                        print_colored(
                            f"NO REGISTER: {RO_number}[{index}]/{row[CAR_NUMBER].value}", "yellow")
                        raise e
                    order = register.all_orders[index]
                    if receipt_number:
                        assert order.receipt_number == receipt_number
                    order.incentive_paid_date = date
                    order.incentive_paid = True
                    order.save()
                    INCENTIVED_SET.add(get_incentive_data(
                        sheet_index, row_index, row, date))
        print_colored("----- INCENTIVED ORDER DATA -----", "blue")
        for incentived_order_data in INCENTIVED_SET:
            print_colored(incentived_order_data, "blue")
